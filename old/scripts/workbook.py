#!/usr/bin/env python3

import os
import xlwt
from xlrd import open_workbook
from xlutils.copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class workbook(object):

    def __init__(self, path):
        self.path = path
        self.wb = None
        self.rb = None

        if not os.path.exists(self.path):
            self.wb = Workbook()
            self.wb.save(filename=self.path)
        else:
            print('--> loading existing workbook... ', end='')
            self.wb = load_workbook(self.path)
            print('done')

    def sheets(self):
        return self.wb.worksheets

    def sheet_names(self):
        return self.wb.sheetnames

    def get_sheet_by_name(self, name):
        idx = 0
        for sheet in self.sheet_names():
            if sheet == name:
                return self.wb.worksheets[idx]
            idx += 1
        return None

    def add_sheet(self, sheet_name):

        # calculate the sheet name if necessary
        existing_sheets = self.sheet_names()
        orig = sheet_name
        i = 2
        while sheet_name in existing_sheets:
            sheet_name = orig + '_%d' % i
            i += 1

        # add the sheet
        self.wb.create_sheet(sheet_name)

        return sheet_name

    def write_pandas(self, sheetname, df, comments=[]):

        # get this sheet
        sheet = self.get_sheet_by_name(sheetname)
        if sheet is None:
            sheet = self.add_sheet(sheetname)

#        for comment in comments:
#            sheet.append(comment)

        for r in dataframe_to_rows(df, index=True, header=True):
            sheet.append(r)

    def write_column(self, start_row, col, sheetname, data):

        # increment row and col indices b/c excel expects 1-based arrays
        start_row += 1
        col += 1

        # get this sheet
        sheet = self.get_sheet_by_name(sheetname)
        if sheet is None:
            sheet = self.add_sheet(sheetname)
#        import pdb; pdb.set_trace()

        # switch to read-only mode to boost performance
#        self.wb = Workbook(write_only=True)
        # write the data
        row = start_row
        for d in data:
            try:
                sheet.cell(row=row, column=col).value = d
            except Exception as e:
                print('failed to write element: row %d, col %d, '
                        'sheet %s, value %s): %s' %
                      (row, col, sheetname, d, e))
            row += 1

    def save(self):
        print('--> saving to xlsx...', end='')
        self.wb.save(self.path)
        print('done')

    def __del__(self):
        self.save()


class workbook_old(object):

    def __init__(self, path):
        self.path = path
        self.wb = None
        self.rb = None

        if not os.path.exists(self.path):
            wb = xlwt.Workbook()
            wb.add_sheet('__home__')
            wb.save(self.path)
        self.rb = open_workbook(self.path)
        self.wb = copy(self.rb)

    def sheets(self):
        self.rb = open_workbook(self.path)
        return self.rb.sheets()

    def sheet_names(self):
        return [s.name for s in self.sheets()]

    def get_sheet_by_name(self, name):
        idx = 0
        for sheet in self.sheets():
            if sheet.name == name:
                return self.wb.get_sheet(idx)
            idx += 1
        return None

    def add_sheet(self, sheet_name):

        # calculate the sheet name if necessary
        existing_sheets = self.sheet_names()
        orig = sheet_name
        i = 2
        while sheet_name in existing_sheets:
            sheet_name = orig + '_%d' % i
            i += 1

        # add the sheet
        self.wb.add_sheet(sheet_name)
        self.save()

        return sheet_name

    def write_column(self, start_row, col, sheetname, data):
        # get this sheet
        sheet = self.get_sheet_by_name(sheetname)
        if sheet is None:
            sheet = self.add_sheet(sheetname)

        # write the data
        row = start_row
        for d in data:
            try:
                sheet.write(row, col, d)
            except Exception as e:
                import pdb; pdb.set_trace()
                print('failed to write element: row %d, col %d, '
                      'sheet %s, value %s)' %
                      (row, col, sheetname, d))
            row += 1

        self.save()

    def save(self):
        self.wb.save(self.path)

